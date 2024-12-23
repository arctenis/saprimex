"""
Application Tkinter pour traiter et formater des fichiers Excel
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

TEMP_FILE = "temp.xlsx"
HEADERS = [
    "TYPE",
    "Raison C/F",
    "Date",
    "Lot",
    "Désignation",
    "Poids",
    "UN",
    "PU",
    "Résultat",
]
YELLOW = "FFFF00"
RED = "FF0000"
LIGHTBLUE = "A0D0FF"
GRAY = "808080"
DARKGRAY = "B0B0B0"
LIGHTGRAY = "E0E0E0"
WHITE = "FFFFFF"
BLACK = "000000"

COULEUR_ENTETE = WHITE
COULEUR_VENTE = RED
COULEUR_CORBEILLE = LIGHTBLUE
COULEUR_FOND_RES_POSITIF = LIGHTGRAY
COULEUR_FOND_RES_NEGATIF = YELLOW
COULEUR_FOND_SOUS_TOTAL = DARKGRAY
COULEUR_FOND_TOTAL = BLACK
COULEUR_POLICE_RES_POSITIF = RED
COULEUR_POLICE_RES_NEGATIF = YELLOW
COULEUR_POLICE_SOUS_TOTAL = BLACK
COULEUR_POLICE_TOTAL = WHITE

output_dir = "G:\\AGNES\\MARGES\\MARGES_PAR_ARRIVAGES"


def process_excel_file():
    """
    Fonction principale appelée lorsque le bouton "Ouvrir" est cliqué.
    """
    excel_file = open_file()
    df = clean_dataframe(excel_file)
    buyer_groups = process_dataframe(df)
    wb, yyyymmdd = create_excel_file(buyer_groups)
    wb = apply_styles(wb)
    final_file_name = f"MARGES PAR ARRIVAGES {yyyymmdd}.xlsx"
    if not os.path.exists(output_dir):
        show_warning("Veuillez sélectionner un répertoire valide.")
        return
    else:
        wb.save(os.path.join(output_dir, final_file_name))
    messagebox.showinfo("", "Fichier traité et créé.")


def apply_styles(wb) -> Workbook:
    """
    Vérifie si un mot-clé est présent sur chacune des lignes et applique des styles spécifiques.
    Efface les colonnes temporaires.
    """
    ws = wb.active
    # thin = Side(border_style="thin", color="000000")
    date_cell = ws["A1"]
    date_cell.font = Font(bold=True, color=RED)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "TYPE":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=GRAY)
                    cell.font = Font(bold=True, color=COULEUR_ENTETE)
            if cell.value == "VENTE":
                for cell in row:
                    cell.font = Font(color=COULEUR_VENTE)
            if cell.value == "Corbeille":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=COULEUR_CORBEILLE)
            # if cell.value == "negative":
            #     for cell in row:
            #         cell.fill = PatternFill("solid", fgColor=YELLOW)
            #         cell.font = Font(color="FF0000", bold=True)
            if cell.value == "neg_result":
                row[8].fill = PatternFill("solid", fgColor=COULEUR_FOND_RES_NEGATIF)
                row[8].font = Font(color=COULEUR_POLICE_RES_NEGATIF, bold=True)
            if cell.value == "Sous-total":
                for cell in row:
                    cell.fill = PatternFill(fgColor=COULEUR_FOND_SOUS_TOTAL)
                    cell.font = Font(color=COULEUR_POLICE_SOUS_TOTAL)
            if cell.value == "Total":
                for cell in row:
                    cell.fill = PatternFill(fgColor=COULEUR_FOND_TOTAL)
                    cell.font = Font(bold=True, color=COULEUR_POLICE_TOTAL)
            # cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Supprimer les colonnes inutiles
    ws.delete_cols(10, amount=4)
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Lettre de la colonne (A, B, C, ...)

        # Calculer la largeur maximale dans chaque colonne (pour chaque cellule)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Ajuster la largeur de la colonne
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width
    return wb


def create_excel_file(buyer_groups):
    wb = Workbook()
    ws = wb.active
    for group in buyer_groups:
        group_total = 0
        ws.append(HEADERS)
        for lot in group:
            group_total += lot["result_total"]
            for r in dataframe_to_rows(lot["df"], index=False, header=False):
                if lot["result_total"] < 0:
                    r.append("negative")
                if r[0] == "TYPE":
                    r.append("header")
                if r[1] == "Corbeille":
                    r.append("recyclebin")
                ws.append(r)
            if lot["result_total"] >= 0.0:
                ws.append(
                    [
                        "Sous-total",
                        "",
                        "",
                        "",
                        "",
                        lot["weight_total"],
                        "",
                        "",
                        lot["result_total"],
                    ]
                )
            else:
                ws.append(
                    [
                        "Sous-total",
                        "",
                        "",
                        "",
                        "",
                        lot["weight_total"],
                        "",
                        "",
                        lot["result_total"],
                        "neg_result",
                    ]
                )
        ws.append(["Total", "", "", "", "", "", "", "", round(group_total, 2)])
        ws.append([])

    yyyymmdd = int(ws["C3"].value)
    ddmmyy = convert_date(yyyymmdd)
    ws.insert_rows(1)
    ws["A1"].value = f"ARRIVAGES DU {ddmmyy}"

    return wb, yyyymmdd


def convert_date(yyyymmdd: int) -> str:
    """
    Convert a date in YYYYMMDD format to DD/MM/YY
    """
    old_date = str(yyyymmdd)
    ddmmyy = old_date[6:8] + "/" + old_date[4:6] + "/" + old_date[2:4]
    return ddmmyy


def process_dataframe(df):
    # Ajouter une colonne temporaire pour stocker les formats de ligne
    df["row_format"] = ""
    # Diviser les données par lot
    lot_df = divide_by_lot(df)
    # Calculer les sous-totaux
    lots = [calculate_subtotals(lot_df) for lot_df in lot_df]
    # Grouper les lots par vendeur ("Raison C/F"/"TYPE")
    # "TYPE" doit être "ACHAT"
    buyers = group_by_buyers(lots)
    return buyers


def group_by_buyers(lots):
    """
    Chaque lot a un vendeur ("Raison C/F"/"TYPE")
    Groupe les lots par vendeur et ajoute la liste des lots à buyer_groups
    """
    buyer_groups = []
    current_buyer = lots[0]["df"]["Raison C/F"][0]
    current_buyer_group = []
    for lot in lots:
        if lot["df"]["Raison C/F"][0] == current_buyer:
            current_buyer_group.append(lot)
        else:
            current_buyer = lot["df"]["Raison C/F"][0]
            buyer_groups.append(current_buyer_group)
            current_buyer_group = [lot]

    return buyer_groups


def divide_by_lot(df):
    """
    Renvoie une liste de DataFrames, un pour chaque lot.
    Seulement les 11 premières lettres du numéro de lot sont considérées.
    """
    lot_dataframes = []
    # Seulement les 11 premières lettres du numéro de lot sont considérées
    current_lot_number = str(df["Lot"][0][:11])
    # Initialiser la dataframe courante avec les mêmes colonnes que la dataframe originale
    current_lot_dataframe = pd.DataFrame(columns=df.columns)
    for _, row in df.iterrows():
        if row["TYPE"] == "ACHAT" or row["TYPE"] == "VENTE":
            if str(row["Lot"])[:11] == current_lot_number:
                current_lot_dataframe = pd.concat(
                    [current_lot_dataframe, row.to_frame().T], ignore_index=True
                )
            else:
                lot_dataframes.append(current_lot_dataframe)
                current_lot_dataframe = pd.DataFrame(columns=df.columns)
                current_lot_dataframe = pd.concat(
                    [current_lot_dataframe, row.to_frame().T], ignore_index=True
                )
                current_lot_number = str(row["Lot"])[:11]
    lot_dataframes.append(current_lot_dataframe)
    return lot_dataframes


def calculate_subtotals(lot_df: pd.DataFrame):
    """
    Ajoute une nouvelle ligne après chaque lot avec le sous-total de "Poids" et "Résultat"
    """
    lot_df["Poids"] = lot_df["Poids"].astype(float)
    lot_df["Résultat"] = lot_df["Résultat"].astype(float)
    weight_total = round(lot_df["Poids"].sum(), 2)
    result_total = round(lot_df["Résultat"].sum(), 2)
    lot = {
        "df": lot_df,
        "weight_total": weight_total,
        "result_total": result_total,
        "negative": result_total < 0,
    }
    return lot


def clean_dataframe(excel_file: str):
    """Supprime les doublons et les colonnes/lignes inutiles du DataFrame."""
    df = pd.read_excel(excel_file)
    df.drop_duplicates(inplace=True)
    df.sort_values(by=["Lot", "TYPE"], inplace=True)
    df = df[df["Code C/F"] != "-REGUL"]
    df.drop(
        columns=["Code C/F", "Commande", "Article", "Colis", "Pièces"], inplace=True
    )
    # Supprimer l'index du DataFrame
    df.reset_index(drop=True, inplace=True)
    if df.empty:
        raise ValueError("Le fichier Excel est vide")
    return df


def open_file() -> str:
    try:
        file_path = filedialog.askopenfilename(
            filetypes=[("Fichier Excel", "*.xlsx *.xls")],
            title="Sélectionnez un fichier",
        )
        if not file_path:
            print("Aucun fichier sélectionné")
        return file_path
    except Exception as e:
        raise ValueError(
            f"Une erreur est survenue lors de la sélection du fichier: {e}"
        )


def select_output_directory():
    """
    Permet à l'utilisateur de sélectionner un répertoire pour le fichier final.
    """
    global output_dir
    try:
        selected_dir = filedialog.askdirectory(title="Sélectionnez un répertoire")
        if not selected_dir:
            show_warning("Aucun répertoire sélectionné.")
            return
        output_dir = selected_dir
        directory_label.config(text=f"{output_dir}")
        return output_dir
    except Exception as e:
        raise ValueError(
            f"Une erreur est survenue lors de la sélection du répertoire: {e}"
        )


def run():
    root = tk.Tk()
    root.title("Saprimex")
    root.geometry("400x400")
    root.resizable(False, False)

    logo = tk.PhotoImage(file="logo.png")
    logo_label = tk.Label(root, image=logo)
    logo_label.pack(pady=20)

    open_button = tk.Button(root, text="Ouvrir", command=process_excel_file)
    open_button.pack(pady=20)

    output_dir_button = tk.Button(
        root, text="Sélectionner un répertoire", command=select_output_directory
    )
    output_dir_button.pack(pady=20)

    select_label = tk.Label(root, text="Répertoire sélectionné :")
    select_label.pack(pady=10)

    global directory_label
    directory_label = tk.Label(root, text=f"{output_dir}")
    directory_label.pack(pady=10)

    root.mainloop()


def show_warning(msg):
    messagebox.showwarning("Attention", msg)


if __name__ == "__main__":
    run()
